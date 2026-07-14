from flask import Flask, jsonify, request, Response
from werkzeug.exceptions import RequestEntityTooLarge
import json
import base64
import io
import time
import traceback
import uuid
from flask_cors import CORS
from werkzeug.utils import secure_filename
import os
import sys
import pandas as pd

# utils 경로 추가
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from utils.vil_processor import process_vil_data, parse_target_current_from_filename, parse_duty_from_filename
from utils.osc_processor import (
    process_osc_data,
    process_lowfreq_pair,
    parse_frequency_duty,
    get_preview_data,
    get_lowfreq_preview_data,
    get_pair_key,
    is_rise_filename,
    is_decay_filename,
)
from utils.master_processor import process_master
from utils.trel_common import parse_minutes_from_filename
from utils.trel_analysis import (
    analyze_single_file,
    DEFAULT_TANGENT_WINDOW_POINTS,
    format_rise_analysis_mode,
    get_preview_data as get_trel_preview,
    parse_after_duty_from_filename,
    parse_vil_processed_for_voltage,
    parse_vil_processed_for_voltage_luminance,
)

app = Flask(__name__)
CORS(app)

TREL_CACHE_TTL_SECONDS = 60 * 30
PROCESSED_TREL_CACHE = {}
PROCESSED_VIL_CACHE = {}
TREL_BATCH_PROGRESS = {'active': False, 'current': 0, 'total': 0, 'filename': '', 'stage': ''}


def parse_tangent_window_points(raw_value) -> int:
    try:
        value = int(float(raw_value))
    except (TypeError, ValueError):
        value = DEFAULT_TANGENT_WINDOW_POINTS
    return max(3, value)


def prune_cache(cache_store):
    now = time.time()
    expired_keys = [
        key for key, value in cache_store.items()
        if now - value['created_at'] > TREL_CACHE_TTL_SECONDS
    ]
    for key in expired_keys:
        cache_store.pop(key, None)


def store_csv_cache(cache_store, payload):
    prune_cache(cache_store)
    cache_key = uuid.uuid4().hex
    cache_store[cache_key] = {
        **payload,
        'created_at': time.time(),
    }
    return cache_key


def csv_text_to_xlsx_bytes(csv_text: str, metadata: dict = None) -> bytes:
    """CSV 문자열을 XLSX 바이너리로 변환. metadata가 있으면 Metadata 시트 추가."""
    try:
        df = pd.read_csv(io.StringIO(csv_text), comment='#')
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Data')
            if metadata:
                pd.DataFrame([metadata]).to_excel(writer, index=False, sheet_name='Metadata')
        return output.getvalue()
    except Exception as e:
        print(f"Excel Conversion Error: {e}", flush=True)
        # Fallback to manual conversion if pandas fails
        import csv
        import openpyxl
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet()
        # Fallback: 메타데이터 라인('#')은 건너뛰고 데이터만 처리 시도
        lines = csv_text.splitlines()
        data_lines = [line for line in lines if not line.startswith('#')]
        
        reader = csv.reader(data_lines)
        for row in reader:
            ws.append(row)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()


def parse_vil_uploaded_for_voltage(file_storage):
    """업로드된 VIL_processed 파일(CSV/XLSX)에서 (time_min, voltage) 목록 추출."""
    data = parse_vil_uploaded_for_voltage_luminance(file_storage)
    return [(t, v) for t, v, _ in data] if data else []


def load_vil_uploaded_for_analysis(file_storage):
    """VIL_processed 업로드 파일 1회 읽기: (data, metadata)."""
    filename = file_storage.filename or ''
    raw = file_storage.read()
    meta = {
        'time_shift_min': 0.0,
        'duty_fraction': parse_duty_from_filename(filename) or 1.0,
        'filename': filename,
    }

    if filename.lower().endswith('.xlsx'):
        try:
            xls = pd.ExcelFile(io.BytesIO(raw))
            if 'Metadata' in xls.sheet_names:
                mdf = pd.read_excel(xls, sheet_name='Metadata')
                if len(mdf) > 0:
                    row = mdf.iloc[0]
                    if 'time_shift_min' in mdf.columns and pd.notna(row.get('time_shift_min')):
                        meta['time_shift_min'] = float(row['time_shift_min'])
                    elif 'time_shift_s' in mdf.columns and pd.notna(row.get('time_shift_s')):
                        meta['time_shift_min'] = float(row['time_shift_s']) / 60.0
                    if 'duty_fraction' in mdf.columns and pd.notna(row.get('duty_fraction')):
                        meta['duty_fraction'] = float(row['duty_fraction'])
            data_sheet = 'Data' if 'Data' in xls.sheet_names else xls.sheet_names[0]
            df = pd.read_excel(xls, sheet_name=data_sheet)
            df.columns = df.columns.str.strip()
            time_col = next((c for c in df.columns if 'Time (min)' in c), None)
            volt_col = next((c for c in df.columns if 'Voltage' in c), None)
            lum_col = next((c for c in df.columns if 'Relative luminance' in c or 'Luminance' in c), None)
            if time_col and volt_col and lum_col:
                return list(zip(df[time_col], df[volt_col], df[lum_col])), meta
            if time_col and volt_col:
                return [(t, v, float('nan')) for t, v in zip(df[time_col], df[volt_col])], meta
            if len(df.columns) >= 3:
                df = df.apply(pd.to_numeric, errors='coerce').dropna()
                return list(zip(df.iloc[:, 0], df.iloc[:, 2], df.iloc[:, 1])), meta
            if len(df.columns) >= 2:
                df = df.apply(pd.to_numeric, errors='coerce').dropna()
                return [(t, v, float('nan')) for t, v in zip(df.iloc[:, 0], df.iloc[:, 1])], meta
        except Exception:
            pass
        return [], meta

    content = raw.decode('utf-8', errors='replace')
    time_shift_s, data = parse_vil_processed_for_voltage_luminance(content)
    if time_shift_s is not None:
        meta['time_shift_min'] = time_shift_s / 60.0
    return data, meta


def parse_vil_uploaded_for_voltage_luminance(file_storage):
    """업로드된 VIL_processed 파일(CSV/XLSX)에서 (time_min, voltage, relative_luminance) 목록 추출."""
    data, _ = load_vil_uploaded_for_analysis(file_storage)
    return data


def parse_vil_uploaded_metadata(file_storage) -> dict:
    """VIL_processed 업로드 파일에서 analysis용 메타데이터 추출."""
    _, meta = load_vil_uploaded_for_analysis(file_storage)
    return meta

# 요청 크기 제한
# MAX_UPLOAD_MB <= 0 이면 Flask 업로드 크기 제한을 비활성화한다(무제한).
MAX_UPLOAD_MB = int(os.environ.get('MAX_UPLOAD_MB', '0'))
if MAX_UPLOAD_MB > 0:
    app.config['MAX_CONTENT_LENGTH'] = MAX_UPLOAD_MB * 1024 * 1024

# Werkzeug 폼 파서 제한 해제
# - MAX_FORM_MEMORY_SIZE: multipart의 non-file 필드 메모리 제한
# - MAX_FORM_PARTS: multipart 파트(파일 개수 포함) 제한
app.config['MAX_FORM_MEMORY_SIZE'] = None
app.config['MAX_FORM_PARTS'] = None


# 413 에러 핸들러
@app.errorhandler(RequestEntityTooLarge)
def handle_request_entity_too_large(e):
    limit_text = f'{MAX_UPLOAD_MB}MB' if MAX_UPLOAD_MB > 0 else '무제한 설정(상위 프록시/서버 제한 가능)'
    return jsonify({
        'success': False,
        'error': f'요청 크기가 너무 큽니다. 전송하려는 파일이 너무 많거나 크기 때문일 수 있습니다. (현재 제한: {limit_text})'
    }), 413


@app.route('/')
def health_root():
    return jsonify({'status': 'ok', 'message': 'Flask backend is running'})


@app.route('/api/health')
def health():
    return jsonify({'status': 'ok', 'message': 'Flask backend is running'})


@app.route('/api/process-vil', methods=['POST'])
def process_vil():
    """
    VIL 파일 처리 API
    - files: VIL이 포함된 CSV 파일들 (multipart/form-data)
    - 또는 file: 단일 파일
    """
    try:
        files = request.files.getlist('files') or ([request.files.get('file')] if request.files.get('file') else [])
        files = [f for f in files if f and f.filename and not f.filename.startswith('._')]
        paths = request.form.getlist('paths') or []

        if not files:
            return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400

        _dam = request.form.get('device_area_mm2', '').strip()
        device_area_mm2 = float(_dam) if _dam else None

        results = []
        processed_paths = set()
        for i, f in enumerate(files):
            filename = secure_filename(f.filename) or f.filename
            rel_path = paths[i] if i < len(paths) else filename

            if 'VIL' not in filename.upper():
                continue
            if rel_path in processed_paths:
                continue
            processed_paths.add(rel_path)

            target = parse_target_current_from_filename(filename)
            if target is None:
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'success': False,
                    'error': '파일명에서 목표 전류를 추출할 수 없습니다.'
                })
                continue

            try:
                # R_shunt, R_osc 파라미터 받기 (기본값 사용)
                r_shunt = request.form.get('r_shunt')
                r_osc = request.form.get('r_osc')
                r_shunt = float(r_shunt) if r_shunt else None
                r_osc = float(r_osc) if r_osc else None
                
                content = f.read().decode('utf-8', errors='replace')
                csv_out, time_shift, meta = process_vil_data(
                    content, target, filename, r_shunt, r_osc, device_area_mm2=device_area_mm2
                )
                xlsx_meta = {
                    'time_shift_s': meta['time_shift_s'],
                    'time_shift_min': meta['time_shift_min'],
                    'duty_fraction': meta.get('duty_fraction', parse_duty_from_filename(filename) or 1.0),
                    'filename': filename,
                }
                xlsx_bytes = csv_text_to_xlsx_bytes(csv_out, metadata=xlsx_meta)
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'output_filename': meta.get('output_filename', filename.replace('.csv', '_processed.xlsx')),
                    'success': True,
                    'csv': csv_out,
                    'cache_key': store_csv_cache(PROCESSED_VIL_CACHE, {
                        'filename': filename,
                        'csv': csv_out,
                        'time_shift_min': meta['time_shift_min'],
                    }),
                    'xlsx_b64': base64.b64encode(xlsx_bytes).decode('ascii'),
                    'time_shift_s': meta['time_shift_s'],
                    'time_shift_min': meta['time_shift_min'],
                    'target_current_ua': meta['target_current_ua'],
                    'original_points': meta['original_points'],
                    'filtered_points': meta['filtered_points']
                })
            except Exception as e:
                results.append({
                    'filename': filename,
                    'relPath': rel_path,
                    'success': False,
                    'error': str(e)
                })

        return jsonify({'success': True, 'results': results})
    except Exception as e:
        print(f"[process-vil] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/preview-osc', methods=['POST'])
def preview_osc():
    """
    오실로스코프 decay 파일 미리보기 (low-freq)
    - sat/zero 구간 지정을 위해 time_ns, CH1 데이터 반환
    """
    try:
        f = request.files.get('file')
        if not f or not f.filename or f.filename.startswith('._'):
            return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400

        content = f.read().decode('utf-8', errors='replace')
        preview = get_lowfreq_preview_data(content)

        if 'error' in preview:
            return jsonify({'success': False, 'error': preview['error']}), 400

        freq, duty = parse_frequency_duty(f.filename or '')
        return jsonify({
            'success': True,
            'filename': f.filename,
            'time_ns': preview['time_ns'],
            'ch1': preview['ch1'],
            'ch2': preview['ch2'],
            'n_points': preview['n_points'],
            'frequency_hz': freq,
            'duty_fraction': duty,
            'mode': 'lowfreq',
        })
    except Exception as e:
        print(f"[preview-osc] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/process-osc', methods=['POST'])
def process_osc():
    """
    Low-frequency rise/decay 페어 TrEL 처리
    - rise_file, decay_file (단일 요청) 또는 pair_key 기반 일괄 처리
    - baseline_start_ns / baseline_end_ns: zero 기준 구간 (decay 후)
    - norm_start_ns / norm_end_ns: saturation 기준 구간 (decay 포화)
    """
    try:
        baseline_start = float(request.form.get('baseline_start_ns', 100000))
        baseline_end = float(request.form.get('baseline_end_ns', 250000))
        norm_start = request.form.get('norm_start_ns', '').strip()
        norm_end = request.form.get('norm_end_ns', '').strip()
        sat_start_ns = float(norm_start) if norm_start else -250000
        sat_end_ns = float(norm_end) if norm_end else -50000

        _dam = request.form.get('device_area_mm2', '').strip()
        device_area_mm2 = float(_dam) if _dam else None

        r_shunt = request.form.get('r_shunt')
        r_osc = request.form.get('r_osc')
        r_shunt = float(r_shunt) if r_shunt else None
        r_osc = float(r_osc) if r_osc else None

        rise_file = request.files.get('rise_file')
        decay_file = request.files.get('decay_file')
        pair_key = request.form.get('pair_key', '').strip()
        rise_path = request.form.get('rise_path', '').strip()
        decay_path = request.form.get('decay_path', '').strip()

        if rise_file and decay_file and rise_file.filename and decay_file.filename:
            pairs = [{
                'rise_file': rise_file,
                'decay_file': decay_file,
                'rise_filename': secure_filename(rise_file.filename) or rise_file.filename,
                'decay_filename': secure_filename(decay_file.filename) or decay_file.filename,
                'rel_path': rise_path or decay_path or pair_key,
                'pair_key': pair_key or get_pair_key(rise_file.filename) or get_pair_key(decay_file.filename),
            }]
        else:
            return jsonify({
                'success': False,
                'error': 'rise_file과 decay_file이 모두 필요합니다.',
            }), 400

        results = []
        for pair in pairs:
            rise_fn = pair['rise_filename']
            decay_fn = pair['decay_filename']
            rel_path = pair['rel_path']
            key = pair['pair_key']

            if not is_rise_filename(rise_fn):
                results.append({
                    'filename': rise_fn,
                    'relPath': rel_path,
                    'success': False,
                    'error': f'rise 파일이 아닙니다: {rise_fn}',
                })
                continue
            if not is_decay_filename(decay_fn):
                results.append({
                    'filename': decay_fn,
                    'relPath': rel_path,
                    'success': False,
                    'error': f'decay 파일이 아닙니다: {decay_fn}',
                })
                continue

            rise_key = get_pair_key(rise_fn)
            decay_key = get_pair_key(decay_fn)
            if rise_key != decay_key:
                results.append({
                    'filename': f'{rise_fn} + {decay_fn}',
                    'relPath': rel_path,
                    'success': False,
                    'error': f'rise/decay 페어 키 불일치: {rise_key} vs {decay_key}',
                })
                continue

            filename_base = key or rise_key or rise_fn.replace('.csv', '')

            try:
                rise_content = pair['rise_file'].read().decode('utf-8', errors='replace')
                decay_content = pair['decay_file'].read().decode('utf-8', errors='replace')
                csv_out, meta = process_lowfreq_pair(
                    rise_content,
                    decay_content,
                    zero_start_ns=baseline_start,
                    zero_end_ns=baseline_end,
                    sat_start_ns=sat_start_ns,
                    sat_end_ns=sat_end_ns,
                    filename_base=filename_base,
                    r_shunt=r_shunt,
                    r_osc=r_osc,
                    device_area_mm2=device_area_mm2,
                )
                results.append({
                    'filename': filename_base,
                    'rise_filename': rise_fn,
                    'decay_filename': decay_fn,
                    'relPath': rel_path,
                    'output_filename': meta['output_filename'],
                    'success': True,
                    'csv': csv_out,
                    'cache_key': store_csv_cache(PROCESSED_TREL_CACHE, {
                        'filename': meta['output_filename'],
                        'csv': csv_out,
                    }),
                    'original_points': meta['original_points'],
                    'format': 'lowfreq',
                })
            except Exception as e:
                results.append({
                    'filename': filename_base,
                    'rise_filename': rise_fn,
                    'decay_filename': decay_fn,
                    'relPath': rel_path,
                    'success': False,
                    'error': str(e),
                })

        return jsonify({'success': True, 'results': results})
    except Exception as e:
        print(f"[process-osc] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/create-master', methods=['POST'])
def create_master():
    """
    VIL 기반 마스터 CSV 생성
    - vil_csv: VIL 처리된 CSV (Time min, Relative luminance)
    - vil_time_shift_min: VIL time shift (분)
    - files: *_TrEL.csv 파일들
    """
    try:
        print('[create-master] 요청 수신', flush=True)
        vil_cache_key = request.form.get('vil_cache_key', '').strip()
        vil_csv = request.form.get('vil_csv', '').strip()
        vil_filename = request.form.get('vil_filename', '').strip()
        vil_time_shift_min_str = request.form.get('vil_time_shift_min', '0').strip()
        vil_time_shift_min = float(vil_time_shift_min_str) if vil_time_shift_min_str else 0.0

        if vil_cache_key:
            prune_cache(PROCESSED_VIL_CACHE)
            cached_vil = PROCESSED_VIL_CACHE.get(vil_cache_key)
            if cached_vil:
                vil_csv = cached_vil['csv']
                vil_time_shift_min = cached_vil['time_shift_min']
                if not vil_filename:
                    vil_filename = cached_vil.get('filename', '')
                print(f'[create-master] VIL cache hit: {vil_cache_key}', flush=True)
            else:
                print(f'[create-master] VIL cache miss: {vil_cache_key}', flush=True)

        master_percents_str = request.form.get('master_percents', '100,90,80,70,60,50').strip()
        try:
            master_percents = [int(x.strip()) for x in master_percents_str.split(',') if x.strip()]
        except ValueError:
            master_percents = [100, 90, 80, 70, 60, 50]
        if not master_percents:
            master_percents = [100, 90, 80, 70, 60, 50]

        cache_keys = [key for key in request.form.getlist('cache_keys') if key]
        files = request.files.getlist('files') or ([request.files.get('file')] if request.files.get('file') else [])
        files = [f for f in files if f and f.filename and not f.filename.startswith('._') and '_TrEL.csv' in f.filename]

        if not vil_csv:
            return jsonify({'success': False, 'error': 'VIL 처리된 CSV가 필요합니다. 마스터 생성은 VIL 데이터를 기반으로 합니다.'}), 400
        if not files and not cache_keys:
            return jsonify({'success': False, 'error': 'TrEL 처리된 CSV 파일이 없습니다. (_TrEL.csv)'}), 400

        print(f'[create-master] 업로드된 TrEL 후보 files={len(files)}, cache_keys={len(cache_keys)}개', flush=True)
        files_data = []
        prune_cache(PROCESSED_TREL_CACHE)
        for cache_key in cache_keys:
            cached = PROCESSED_TREL_CACHE.get(cache_key)
            if not cached:
                print(f'[create-master] TrEL cache miss: {cache_key}', flush=True)
                continue
            files_data.append((cached['filename'], cached['csv']))
            parsed_minutes = parse_minutes_from_filename(cached['filename'])
            print(
                f'[create-master] 캐시 후보 파일: {cached["filename"]} | parsed_minutes={parsed_minutes}',
                flush=True,
            )

        for f in files:
            content = f.read().decode('utf-8', errors='replace')
            files_data.append((f.filename, content))
            parsed_minutes = parse_minutes_from_filename(f.filename)
            print(
                f'[create-master] 후보 파일: {f.filename} | parsed_minutes={parsed_minutes}',
                flush=True,
            )

        if not files_data:
            return jsonify({'success': False, 'error': '마스터 생성용 TrEL 캐시가 만료되었거나 읽을 수 없습니다. 다시 처리 후 시도해주세요.'}), 400

        print(
            f'[create-master] master_percents={master_percents}, vil_time_shift_min={vil_time_shift_min}',
            flush=True,
        )
        print(f'[create-master] TrEL {len(files_data)}개 로드, process_master 시작', flush=True)

        vil_filename = request.form.get('vil_filename', '').strip() or vil_filename
        target_current_ua = parse_target_current_from_filename(vil_filename) if vil_filename else None

        _dam = request.form.get('device_area_mm2', '').strip()
        device_area_mm2 = float(_dam) if _dam else None
        r_shunt = request.form.get('r_shunt')
        r_osc = request.form.get('r_osc')
        r_shunt = float(r_shunt) if r_shunt else None
        r_osc = float(r_osc) if r_osc else None

        xlsx_bytes, _summary, metadata = process_master(
            vil_csv,
            vil_time_shift_min,
            files_data,
            percent_list=master_percents,
            target_current_ua=target_current_ua,
            device_area_mm2=device_area_mm2,
            r_shunt=r_shunt,
            r_osc=r_osc,
            vil_filename=vil_filename,
        )
        output_filename = metadata.get('output_filename', 'TrEL_Master.xlsx')
        print(f'[create-master] metadata={json.dumps(metadata, ensure_ascii=False)}', flush=True)
        print('[create-master] 완료', flush=True)

        # XLSX 파일 + X-Master-Metadata 헤더로 선택 파일 정보 전달
        resp = Response(
            xlsx_bytes,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{output_filename}"',
                'X-Master-Metadata': json.dumps(metadata, ensure_ascii=False),
                'Access-Control-Expose-Headers': 'X-Master-Metadata, Content-Disposition',
            }
        )
        return resp
    except RequestEntityTooLarge as e:
        return handle_request_entity_too_large(e)
    except ImportError as e:
        return jsonify({'success': False, 'error': f'openpyxl 설치 필요: pip install openpyxl'}), 500
    except Exception as e:
        print(f"[create-master] Error: {e}", flush=True)
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/trel-analysis-preview', methods=['POST'])
def trel_analysis_preview():
    """
    TrEL 배치 분석 미리보기 (첫 파일)
    - low_pct, high_pct, n_decay, rise_mode, decay_fit_start_us
    """
    try:
        f = request.files.get('file')
        if not f or not f.filename or f.filename.startswith('._'):
            return jsonify({'success': False, 'error': '파일이 없습니다.'}), 400
        low_pct = float(request.form.get('low_pct', 0.1))
        high_pct = float(request.form.get('high_pct', 99))
        n_decay = int(request.form.get('n_decay', 2))
        rise_mode = request.form.get('rise_mode', 'tangent')
        decay_fit_start_us = float(request.form.get('decay_fit_start_us', 0.0))
        decay_fit_end_raw = request.form.get('decay_fit_end_us', '').strip()
        decay_fit_end_us = float(decay_fit_end_raw) if decay_fit_end_raw else None
        tangent_window_points = parse_tangent_window_points(
            request.form.get('tangent_window_points', DEFAULT_TANGENT_WINDOW_POINTS)
        )
        decay_init_json = request.form.get('decay_initial_params')
        decay_initial_params = None
        if decay_init_json:
            try:
                parsed = json.loads(decay_init_json)
                if isinstance(parsed, list) and len(parsed) >= 3:
                    decay_initial_params = [float(x) for x in parsed]
            except (json.JSONDecodeError, TypeError, ValueError):
                pass
        content = f.read().decode('utf-8', errors='replace')
        preview = get_trel_preview(
            content,
            low_pct,
            high_pct,
            n_decay,
            rise_mode=rise_mode,
            decay_fit_start_us=decay_fit_start_us,
            decay_fit_end_us=decay_fit_end_us,
            decay_initial_params=decay_initial_params,
            tangent_window_points=tangent_window_points,
        )
        if preview.get('error'):
            return jsonify({'success': False, 'error': preview['error']}), 400
        return jsonify({'success': True, 'filename': f.filename, **preview})
    except Exception as e:
        print(f"[trel_analysis_preview] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/trel-analysis-progress', methods=['GET'])
def trel_analysis_progress():
    """TrEL 배치 분석 진행 상황 조회"""
    return jsonify(TREL_BATCH_PROGRESS)


@app.route('/api/trel-analysis-batch', methods=['POST'])
def trel_analysis_batch():
    """
    TrEL 배치 분석 - Rise, Saturation, Decay 추출 후 Excel 저장 (Pandas Optimized)
    """
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return jsonify({'success': False, 'error': 'openpyxl 설치 필요: pip install openpyxl'}), 500

    try:
        files = request.files.getlist('files') or ([request.files.get('file')] if request.files.get('file') else [])
        vil_files = request.files.getlist('vil_files') or []
        files = [f for f in files if f and f.filename and not f.filename.startswith('._') and f.filename.lower().endswith('.csv') and ('_TrEL' in f.filename or 'TrEL' in f.filename) and 'VIL' not in f.filename.upper()]
        vil_files = [f for f in vil_files if f and f.filename and not f.filename.startswith('._') and 'VIL' in f.filename.upper() and '_processed' in f.filename]
        low_pct = float(request.form.get('low_pct', 0.1))
        high_pct = float(request.form.get('high_pct', 99))
        n_decay = int(request.form.get('n_decay', 2))
        rise_mode = request.form.get('rise_mode', 'tangent')
        decay_fit_start_us = float(request.form.get('decay_fit_start_us', 0.0))
        decay_fit_end_raw = request.form.get('decay_fit_end_us', '').strip()
        decay_fit_end_us = float(decay_fit_end_raw) if decay_fit_end_raw else None
        tangent_window_points = parse_tangent_window_points(
            request.form.get('tangent_window_points', DEFAULT_TANGENT_WINDOW_POINTS)
        )

        if not files:
            return jsonify({'success': False, 'error': 'CSV 파일이 없습니다.'}), 400

        TREL_BATCH_PROGRESS['active'] = True
        TREL_BATCH_PROGRESS['total'] = len(files)
        TREL_BATCH_PROGRESS['current'] = 0
        TREL_BATCH_PROGRESS['filename'] = ''
        TREL_BATCH_PROGRESS['stage'] = '파일 로드'

        decay_init_json = request.form.get('decay_initial_params')
        preview_decay_popt = None
        if decay_init_json:
            try:
                preview_decay_popt = json.loads(decay_init_json)
                if not isinstance(preview_decay_popt, list) or len(preview_decay_popt) < 3:
                    preview_decay_popt = None
            except (json.JSONDecodeError, TypeError):
                preview_decay_popt = None

        vil_time_voltage_lum = []
        vil_time_shift_min = 0.0
        vil_duty_fraction = 1.0
        vil_shift_form = request.form.get('vil_time_shift_min', '').strip()
        vil_duty_form = request.form.get('vil_duty_fraction', '').strip()
        if vil_shift_form:
            try:
                vil_time_shift_min = float(vil_shift_form)
            except ValueError:
                pass
        if vil_duty_form:
            try:
                vil_duty_fraction = float(vil_duty_form)
            except ValueError:
                pass

        for vf in vil_files:
            data, vil_meta = load_vil_uploaded_for_analysis(vf)
            if vil_meta.get('time_shift_min'):
                vil_time_shift_min = vil_meta['time_shift_min']
            if vil_meta.get('duty_fraction'):
                vil_duty_fraction = vil_meta['duty_fraction']
            if data:
                vil_time_voltage_lum.extend(data)
        if vil_time_voltage_lum:
            seen = {}
            for t, v, lum in sorted(vil_time_voltage_lum, key=lambda x: x[0]):
                if t not in seen:
                    seen[t] = (v, lum)
            vil_time_voltage_lum = [(t, seen[t][0], seen[t][1]) for t in sorted(seen.keys())]

        # 시간 순 정렬 (time_min 또는 after_duty)
        def _file_sort_key(file_obj):
            tm = parse_minutes_from_filename(file_obj.filename)
            ad = parse_after_duty_from_filename(file_obj.filename) or ''
            return (tm is None, tm if tm is not None else float('inf'), ad)

        files_sorted = sorted(files, key=_file_sort_key)
        # 재읽기 위해 content 저장 (file stream은 1회만 읽기 가능)
        items = []
        for f in files_sorted:
            content = f.read().decode('utf-8', errors='replace')
            items.append({'content': content, 'filename': f.filename})

        # 10분에 가장 가까운 파일 인덱스 (미리보기 기준)
        def _time_for_seed(idx):
            tm = parse_minutes_from_filename(items[idx]['filename'])
            return tm if tm is not None else 10.0
        idx_preview = 0
        if len(items) > 1:
            best_diff = abs(_time_for_seed(0) - 10.0)
            for i in range(1, len(items)):
                d = abs(_time_for_seed(i) - 10.0)
                if d < best_diff:
                    best_diff = d
                    idx_preview = i

        results = [None] * len(items)
        total = len(items)

        def _fit(idx, decay_init):
            it = items[idx]
            return analyze_single_file(
                it['content'], it['filename'], low_pct, high_pct, n_decay,
                rise_mode=rise_mode,
                vil_time_voltage=vil_time_voltage_lum if vil_time_voltage_lum else None,
                vil_time_shift_min=vil_time_shift_min,
                vil_duty_fraction=vil_duty_fraction,
                decay_fit_start_us=decay_fit_start_us,
                decay_fit_end_us=decay_fit_end_us,
                decay_initial_params=decay_init,
                tangent_window_points=tangent_window_points,
            )

        def _set_progress(cur, fn, stage):
            TREL_BATCH_PROGRESS['current'] = cur
            TREL_BATCH_PROGRESS['total'] = total
            TREL_BATCH_PROGRESS['filename'] = fn
            TREL_BATCH_PROGRESS['stage'] = stage

        _set_progress(0, '', '준비')
        decay_seed = preview_decay_popt
        R2_MIN_DECAY = 0.85
        for i, idx in enumerate(range(idx_preview, len(items))):
            _set_progress(idx - idx_preview + 1, items[idx]['filename'], '순방향')
            r = _fit(idx, decay_seed)
            results[idx] = r
            decay_r2 = r.get('decay_r2')
            if isinstance(r.get('popt'), list) and (decay_r2 is None or decay_r2 >= R2_MIN_DECAY):
                decay_seed = r.get('popt')

        decay_seed = preview_decay_popt
        for i, idx in enumerate(range(idx_preview - 1, -1, -1)):
            _set_progress(len(items) - idx, items[idx]['filename'], '역방향')
            r = _fit(idx, decay_seed)
            results[idx] = r
            decay_r2 = r.get('decay_r2')
            if isinstance(r.get('popt'), list) and (decay_r2 is None or decay_r2 >= R2_MIN_DECAY):
                decay_seed = r.get('popt')

        has_voltage = any(row.get('voltage') is not None for row in results if row)
        if has_voltage:
            results.sort(key=lambda r: (r.get('time_min') is None, r.get('time_min') or float('inf')))
        else:
            results.sort(key=lambda r: (r.get('after_duty') is None or r.get('after_duty') == '', r.get('after_duty') or ''))

        # Excel 생성 (Pandas) - 분석 실패/빈 결과 행 제외
        excel_rows = []
        first_col_name = 'Time (min)' if has_voltage else 'duty 뒤'

        def _has_meaningful_data(r):
            if not r:
                return False
            if r.get('error'):
                return False
            return (
                r.get('t_delay') is not None
                or r.get('tau_1') is not None
                or r.get('spike_integral') is not None
            )

        for r in results:
            if not _has_meaningful_data(r):
                continue
            row = {}
            row[first_col_name] = r.get('time_min') if has_voltage else r.get('after_duty')
            
            if has_voltage:
                row['Voltage (V)'] = r.get('voltage')
                row['Relative luminance (a.u.)'] = r.get('relative_luminance')
                
            row['t_delay (μs)'] = r.get('t_delay')
            row['t_rise (μs)'] = r.get('t_rise')
            row['t_saturation (μs)'] = r.get('t_saturation')
            
            for i in range(1, n_decay + 1):
                row[f'tau_{i} (μs)'] = r.get(f'tau_{i}')
                row[f'f_{i}'] = r.get(f'f_{i}')
                
            row['tau_avg (μs)'] = r.get('tau_avg')
            row['Spike Integral (nC/cm²)'] = r.get('spike_integral')
            row['Spike Decay Time (μs)'] = r.get('spike_decay_time_us')
            row['Rise Slope (a.u./μs)'] = r.get('rise_slope')
            row['Decay R²'] = r.get('decay_r2')
            excel_rows.append(row)
            
        df = pd.DataFrame(excel_rows)
        
        # 컬럼 순서 지정 (Voltage 오른쪽에 Relative luminance)
        cols = [first_col_name]
        if has_voltage:
            cols.append('Voltage (V)')
            cols.append('Relative luminance (a.u.)')
        cols.extend(['t_delay (μs)', 't_rise (μs)', 't_saturation (μs)'])
        for i in range(1, n_decay + 1):
            cols.extend([f'tau_{i} (μs)', f'f_{i}'])
        cols.extend(['tau_avg (μs)', 'Spike Integral (nC/cm²)', 'Spike Decay Time (μs)', 'Rise Slope (a.u./μs)', 'Decay R²'])
        
        # 존재하는 컬럼만 선택
        cols = [c for c in cols if c in df.columns]
        df = df[cols]

        rise_mode_label = format_rise_analysis_mode(rise_mode)
        sheet_name = f'TrEL Analysis ({rise_mode_label})'
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]
            
            # 헤더 볼드 처리 및 컬럼 너비 조정
            for cell in ws[1]:
                cell.font = Font(bold=True)
                
            for i, col in enumerate(cols, 1):
                w = 18 if i == 1 and not has_voltage else 14
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

        rise_mode_token = rise_mode_label.replace(' ', '')
        fit_start_token = str(decay_fit_start_us).replace('.', 'p')
        output_filename = f'TrEL_Analysis_{rise_mode_token}_fitStart{fit_start_token}us.xlsx'

        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={output_filename}'}
        )
    except Exception as e:
        print(f"[trel_analysis_batch] Error: {e}", flush=True)
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        TREL_BATCH_PROGRESS['active'] = False

if __name__ == '__main__':
    port_str = os.environ.get('PORT')
    if port_str:
        port = int(port_str)
    elif len(sys.argv) > 1:
        port = int(sys.argv[1])
    else:
        port = 8080

    host = os.environ.get('HOST', '0.0.0.0')

    print(f"Flask app starting...")
    print(f"Host: {host}, Port: {port}")
    sys.stdout.flush()

    try:
        print(f"Server started: http://{host}:{port}")
        sys.stdout.flush()
        app.run(debug=False, use_reloader=False, port=port, host=host)
    except OSError as e:
        error_msg = str(e)
        print(f"Port error: {error_msg}")
        sys.stdout.flush()
        if 'Address already in use' in error_msg or 'Port already in use' in error_msg:
            print(f"Port {port} is already in use.")
            print(f"Trying port 5001 instead...")
            sys.stdout.flush()
            app.run(debug=False, use_reloader=False, port=5001, host=host)
        else:
            raise
